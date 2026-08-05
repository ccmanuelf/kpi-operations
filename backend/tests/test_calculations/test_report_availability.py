"""Guards that the report generators use real availability, not placeholders."""

from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from backend.reports.excel_generator import ExcelReportGenerator
from backend.reports.pdf_generator import PDFReportGenerator
from backend.tests.fixtures.factories import TestDataFactory

REPORTS_DIR = Path(__file__).resolve().parents[2] / "reports"


class TestNoPlaceholderAvailability:
    def test_excel_generator_has_no_hardcoded_availability(self):
        src = (REPORTS_DIR / "excel_generator.py").read_text()
        assert '"availability": 90.0' not in src
        assert "calculate_availability_pure" in src

    def test_pdf_generator_has_no_hardcoded_availability(self):
        src = (REPORTS_DIR / "pdf_generator.py").read_text()
        assert "[85.0] * len" not in src
        assert "calculate_availability_pure" in src


def _seed_production_entry(
    db, *, run_time_hours: Decimal, downtime_hours: Decimal, entry_date: date, entry_time: time = time(18, 0)
):
    """Insert one ProductionEntry with known run/downtime hours through real FKs,
    timestamped late in the day (18:00 by default) on entry_date rather than
    midnight. Combined with querying start_date == end_date == entry_date, this
    exercises the DateTime-vs-date upper-bound boundary case (#145 pattern)."""
    client = TestDataFactory.create_client(db)
    user = TestDataFactory.create_user(db, role="admin", client_id=client.client_id)
    product = TestDataFactory.create_product(db, client_id=client.client_id)
    shift = TestDataFactory.create_shift(db, client_id=client.client_id)
    entry = TestDataFactory.create_production_entry(
        db,
        client_id=client.client_id,
        product_id=product.product_id,
        shift_id=shift.shift_id,
        entered_by=user.user_id,
        production_date=entry_date,
        run_time_hours=run_time_hours,
        downtime_hours=downtime_hours,
    )
    entry.production_date = datetime.combine(entry_date, entry_time)
    db.flush()
    return client


class TestAvailabilityValueEquality:
    """Spec §7: feed known run/downtime hours through the REAL fetch path and
    assert the emitted availability figure exactly. run=7h + downtime=1h ->
    scheduled=8h -> availability == 87.5, per calculate_availability_pure's
    (scheduled - downtime) / scheduled * 100 formula.

    The fixture entry is timestamped 18:00 on entry_date and queried with
    start_date == end_date == entry_date — the DateTime-vs-date upper-bound
    boundary case fixed in this change (the #145 `datetime.combine(start_date,
    datetime.min.time())` / `datetime.combine(end_date, datetime.max.time())`
    pattern, applied to every `_fetch_*` method in both generators). An entry
    timestamped anywhere on end_date, not just midnight, must be included.
    """

    def test_excel_fetch_production_data_availability_is_exact(self, transactional_db):
        entry_date = date(2026, 1, 15)
        client = _seed_production_entry(
            transactional_db,
            run_time_hours=Decimal("7.0"),
            downtime_hours=Decimal("1.0"),
            entry_date=entry_date,
        )

        rows = ExcelReportGenerator(transactional_db)._fetch_production_data(client.client_id, entry_date, entry_date)

        assert len(rows) == 1
        assert rows[0]["availability"] == 87.5

    def test_pdf_fetch_kpi_details_availability_is_exact(self, transactional_db):
        entry_date = date(2026, 1, 15)
        client = _seed_production_entry(
            transactional_db,
            run_time_hours=Decimal("7.0"),
            downtime_hours=Decimal("1.0"),
            entry_date=entry_date,
        )

        details = PDFReportGenerator(transactional_db)._fetch_kpi_details(
            "availability", client.client_id, entry_date, entry_date
        )

        assert details["Current Value"] == "87.5%"
        assert details["Average (Period)"] == "87.5%"
        assert details["Best Day"] == "87.5%"
        assert details["Worst Day"] == "87.5%"


def _seed_downtime_entry(
    db, *, client, user, category: str, duration_minutes: int, entry_date: date, entry_time: time = time(12, 0)
):
    """Insert one DowntimeEntry with a known root_cause_category and duration
    through the real factory path (Task 8: downtime-cause-taxonomy)."""
    entry = TestDataFactory.create_downtime_entry(
        db,
        client_id=client.client_id,
        reported_by=user.user_id,
        downtime_reason="EQUIPMENT_FAILURE",
        shift_date=datetime.combine(entry_date, entry_time),
        duration_minutes=duration_minutes,
        root_cause_category=category,
    )
    db.flush()
    return entry


class TestDowntimeSheetByCategorySummary:
    """Task 8 (downtime-cause-taxonomy): the Downtime Analysis sheet gains a
    By-Category rollup (events / total minutes / % of total) inserted above
    the detail table, aggregating the already-fetched _fetch_downtime_data
    rows (see spec §... in docs/superpowers/specs/2026-07-31-downtime-cause-
    taxonomy-design.md)."""

    def test_downtime_sheet_has_by_category_summary_with_exact_values(self, transactional_db):
        entry_date = date(2026, 1, 15)
        client = TestDataFactory.create_client(transactional_db)
        user = TestDataFactory.create_user(transactional_db, role="admin", client_id=client.client_id)
        transactional_db.flush()

        # machine: 90min + 30min = 120min total, 2 events
        _seed_downtime_entry(
            transactional_db, client=client, user=user, category="machine", duration_minutes=90, entry_date=entry_date
        )
        _seed_downtime_entry(
            transactional_db, client=client, user=user, category="machine", duration_minutes=30, entry_date=entry_date
        )
        # materials: 60min, 1 event
        _seed_downtime_entry(
            transactional_db,
            client=client,
            user=user,
            category="materials",
            duration_minutes=60,
            entry_date=entry_date,
        )
        transactional_db.commit()

        wb = Workbook()
        ExcelReportGenerator(transactional_db)._create_downtime_sheet(wb, client.client_id, entry_date, entry_date)
        ws = wb["Downtime Analysis"]

        # Derivation: grand total minutes = 120 (machine) + 60 (materials) = 180.
        # machine % = 120 / 180 * 100 = 66.666...  -> round(., 1) = 66.7
        # materials % = 60 / 180 * 100 = 33.333...  -> round(., 1) = 33.3
        # 66.7 + 33.3 == 100.0 exactly (Cycle 1 minimal rollup, no rounding drift
        # at this precision).
        header_row = [c.value for c in ws[3]][:4]
        machine_row = [c.value for c in ws[4]][:4]
        materials_row = [c.value for c in ws[5]][:4]

        # Finding 4 (final-review): the By-Category column writes a display
        # label ("Machine" / "Materials"), not the raw enum code — backend
        # has no i18n, so title-case is the display convention (spec §7).
        assert header_row == ["By Category", "Events", "Total Minutes", "% of Total"]
        assert machine_row == ["Machine", 2, 120.0, 66.7]
        assert materials_row == ["Materials", 1, 60.0, 33.3]
        assert machine_row[3] + materials_row[3] == 100.0

        # Detail table header must be pushed below the summary block + blank
        # separator row (row 6 blank, row 7 detail header), not overwritten by it.
        detail_header_row = [c.value for c in ws[7]][:6]
        assert detail_header_row == ["Date", "Machine/Line", "Category", "Duration (hrs)", "Impact %", "Root Cause"]

    def test_downtime_sheet_by_category_summary_empty_when_no_data(self, transactional_db):
        """No downtime rows in range -> summary header still renders, no
        category rows, no divide-by-zero (grand total falls back to 1.0)."""
        client = TestDataFactory.create_client(transactional_db)
        transactional_db.commit()

        wb = Workbook()
        ExcelReportGenerator(transactional_db)._create_downtime_sheet(
            wb, client.client_id, date(2026, 1, 15), date(2026, 1, 15)
        )
        ws = wb["Downtime Analysis"]

        assert [c.value for c in ws[3]][:4] == ["By Category", "Events", "Total Minutes", "% of Total"]
        # No category present -> row 4 is the blank separator, not a data row.
        assert [c.value for c in ws[4]][:4] == [None, None, None, None]


class TestExcelGeneratorEmptySheetSelectionGuard:
    """M-4: an empty (or all-unknown-key) sheet selection must fail loudly,
    not silently produce a zero-sheet workbook that crashes openpyxl on save.
    """

    def test_empty_sheet_selection_raises_value_error(self, transactional_db):
        generator = ExcelReportGenerator(transactional_db)

        with pytest.raises(ValueError, match="sheets selected no valid sheet keys"):
            generator.generate_report(
                client_id=None,
                start_date=date(2026, 1, 1),
                end_date=date(2026, 1, 31),
                sheets=["not-a-real-sheet-key"],
            )


def _seed_otd_net_of_justified_scenario(db, client):
    """Seeds the exact Task 5 scenario (test_true_otd_gross_vs_net_with_
    derivation): 5 delivered COMPLETED work orders (2 on-time, 1 late+
    justified/customer_request, 1 late+unjustified, 1 late+unclassified) plus
    1 undelivered past-due order. gross = 2/5*100 = 40.0, net =
    (2+1)/5*100 = 60.0. standard_otd mirrors true_otd exactly here (all 5
    delivered orders are COMPLETED)."""
    from backend.orm.work_order import WorkOrder, WorkOrderStatus

    wo1 = WorkOrder(
        work_order_id="WO-XLNET-001",
        client_id=client.client_id,
        style_model="XLNET-STYLE",
        planned_quantity=10,
        status=WorkOrderStatus.COMPLETED,
        planned_ship_date=datetime(2026, 1, 10, 12, 0),
        actual_delivery_date=datetime(2026, 1, 10, 8, 0),  # before planned -> on time
    )
    wo2 = WorkOrder(
        work_order_id="WO-XLNET-002",
        client_id=client.client_id,
        style_model="XLNET-STYLE",
        planned_quantity=10,
        status=WorkOrderStatus.COMPLETED,
        planned_ship_date=datetime(2026, 1, 15, 12, 0),
        actual_delivery_date=datetime(2026, 1, 14, 8, 0),  # before planned -> on time
    )
    wo3 = WorkOrder(
        work_order_id="WO-XLNET-003",
        client_id=client.client_id,
        style_model="XLNET-STYLE",
        planned_quantity=10,
        status=WorkOrderStatus.COMPLETED,
        planned_ship_date=datetime(2026, 1, 5, 12, 0),
        actual_delivery_date=datetime(2026, 1, 8, 8, 0),  # after planned -> late
        delay_classification="justified",
        justified_delay_reason="customer_request",
    )
    wo4 = WorkOrder(
        work_order_id="WO-XLNET-004",
        client_id=client.client_id,
        style_model="XLNET-STYLE",
        planned_quantity=10,
        status=WorkOrderStatus.COMPLETED,
        planned_ship_date=datetime(2026, 1, 6, 12, 0),
        actual_delivery_date=datetime(2026, 1, 9, 8, 0),  # after planned -> late
        delay_classification="unjustified",
    )
    wo5 = WorkOrder(
        work_order_id="WO-XLNET-005",
        client_id=client.client_id,
        style_model="XLNET-STYLE",
        planned_quantity=10,
        status=WorkOrderStatus.COMPLETED,
        planned_ship_date=datetime(2026, 1, 7, 12, 0),
        actual_delivery_date=datetime(2026, 1, 12, 8, 0),  # after planned -> late
        # delay_classification left NULL -> unclassified
    )
    wo6 = WorkOrder(
        work_order_id="WO-XLNET-006",
        client_id=client.client_id,
        style_model="XLNET-STYLE",
        planned_quantity=10,
        status=WorkOrderStatus.IN_PROGRESS,
        planned_ship_date=datetime(2026, 1, 3, 12, 0),  # before `end` -> is_late(wo, end) True
        actual_delivery_date=None,
    )

    db.add_all([wo1, wo2, wo3, wo4, wo5, wo6])
    db.commit()


class TestExcelSummarySheetOTDGrossAndNet:
    """Cycle 2 Task 7: Executive Summary's KPI table gains OTD gross +
    net-of-justified rows, sourced from calculate_true_otd's standard_otd
    (spec: "the comprehensive Excel report's OTD row shows gross + net";
    added as two rows since this table is one-KPI-per-row with no per-KPI
    secondary-value column). Both rows come from the SAME
    calculate_true_otd(...) call so gross/net stay internally consistent
    within the sheet."""

    def test_fetch_kpi_summary_data_otd_gross_and_net_exact(self, transactional_db):
        client = TestDataFactory.create_client(transactional_db)
        _seed_otd_net_of_justified_scenario(transactional_db, client)

        kpi_data = ExcelReportGenerator(transactional_db)._fetch_kpi_summary_data(
            client.client_id, date(2026, 1, 1), date(2026, 1, 31)
        )

        by_name = {row["name"]: row for row in kpi_data}
        assert by_name["OTD"]["current"] == 40.0
        assert by_name["OTD (Net of Justified)"]["current"] == 60.0

    def test_fetch_kpi_summary_data_otd_omitted_without_client_id(self, transactional_db):
        """All-clients comprehensive report: calculate_true_otd is inherently
        single-client (mirrors /api/kpi/otd omitting true_otd/standard_otd
        for multi/all-client scope) — no OTD row is added, not an error."""
        client = TestDataFactory.create_client(transactional_db)
        _seed_otd_net_of_justified_scenario(transactional_db, client)

        kpi_data = ExcelReportGenerator(transactional_db)._fetch_kpi_summary_data(
            None, date(2026, 1, 1), date(2026, 1, 31)
        )

        names = [row["name"] for row in kpi_data]
        assert "OTD" not in names
        assert "OTD (Net of Justified)" not in names

    def test_summary_sheet_renders_otd_gross_and_net_at_exact_cells(self, transactional_db):
        """Full render through _create_summary_sheet: with no production/
        quality/attendance data seeded, OTD is the only KPI in the table, so
        its two rows land at exact, predictable cell positions (header row 7,
        data starts row 8)."""
        client = TestDataFactory.create_client(transactional_db)
        _seed_otd_net_of_justified_scenario(transactional_db, client)

        wb = Workbook()
        ExcelReportGenerator(transactional_db)._create_summary_sheet(
            wb, client.client_id, date(2026, 1, 1), date(2026, 1, 31)
        )
        ws = wb["Executive Summary"]

        assert ws["A8"].value == "OTD"
        assert ws["B8"].value == 40.0
        assert ws["A9"].value == "OTD (Net of Justified)"
        assert ws["B9"].value == 60.0


class TestExcelSummarySheetOTDZeroDeliveryGuard:
    """A window with zero delivered orders must not render "0.0% / At Risk"
    OTD rows — that's an absence-of-data signal, not a real 0% on-time rate.
    Guards on calculate_true_otd's standard_otd.total (the delivered-order
    denominator both percentage/net_percentage divide by)."""

    def test_fetch_kpi_summary_data_omits_otd_when_zero_delivered(self, transactional_db):
        client = TestDataFactory.create_client(transactional_db)
        transactional_db.commit()
        # No work orders seeded at all -> standard_otd.total == 0.

        kpi_data = ExcelReportGenerator(transactional_db)._fetch_kpi_summary_data(
            client.client_id, date(2026, 1, 1), date(2026, 1, 31)
        )

        names = [row["name"] for row in kpi_data]
        assert "OTD" not in names
        assert "OTD (Net of Justified)" not in names

    def test_summary_sheet_has_no_otd_rows_when_zero_delivered(self, transactional_db):
        """Full render through _create_summary_sheet: with no production/
        quality/attendance/work-order data seeded at all, kpi_data is
        entirely empty, so the table has a header row (7) and no data rows —
        row 8 (where OTD would have landed) stays blank."""
        client = TestDataFactory.create_client(transactional_db)
        transactional_db.commit()

        wb = Workbook()
        ExcelReportGenerator(transactional_db)._create_summary_sheet(
            wb, client.client_id, date(2026, 1, 1), date(2026, 1, 31)
        )
        ws = wb["Executive Summary"]

        assert ws["A7"].value == "KPI"
        assert ws["B7"].value == "Current Value"
        assert ws["A8"].value is None
        assert ws["B8"].value is None

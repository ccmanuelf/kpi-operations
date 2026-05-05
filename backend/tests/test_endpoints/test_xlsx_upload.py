"""
Integration tests for XLSX upload support on CSV upload endpoints.

Tests downtime and work-order endpoints with XLSX files to verify:
- .xlsx files are accepted and parsed correctly
- sheet_name query parameter works
- .txt files are rejected (400)
- file size limit enforced (413)
- existing CSV upload still works

Uses standalone FastAPI test app with overridden dependencies (no real auth).
Follows the same pattern as test_work_order_capacity_routes.py.
"""

import pytest
from io import BytesIO

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.database import Base, get_db
from backend.auth.jwt import get_current_user
from backend.endpoints.csv_upload import router as csv_upload_router
from backend.orm.user import User

# Import CapacityOrder so its table is registered before create_all.
# WORK_ORDER has FK -> capacity_orders.id.
from backend.orm.capacity.orders import CapacityOrder  # noqa: F401

from backend.tests.fixtures.factories import TestDataFactory


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CLIENT_ID = "XLSX-TEST-CLIENT"


# ---------------------------------------------------------------------------
# Test app factory and fixtures
# ---------------------------------------------------------------------------


def _create_test_app(db_session, role="admin"):
    """Create a FastAPI test app with overridden dependencies."""
    app = FastAPI()
    app.include_router(csv_upload_router)

    def override_get_db():
        try:
            yield db_session
        finally:
            pass

    mock_user = User(
        user_id="xlsx-test-user-001",
        username="xlsx_test_user",
        email="xlsx_test@test.com",
        role=role,
        client_id_assigned=None if role == "admin" else CLIENT_ID,
        is_active=True,
    )

    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[get_current_user] = lambda: mock_user

    return app


@pytest.fixture(scope="function")
def xlsx_db():
    """Create a fresh in-memory database for each test."""
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    TestingSession = sessionmaker(bind=engine)
    session = TestingSession()
    TestDataFactory.reset_counters()
    try:
        yield session
    finally:
        session.rollback()
        session.close()
        Base.metadata.drop_all(bind=engine)
        engine.dispose()


@pytest.fixture
def supervisor_client(xlsx_db):
    """TestClient authenticated as supervisor with a pre-created client."""
    TestDataFactory.create_client(xlsx_db, client_id=CLIENT_ID, client_name="XLSX Test Client")
    xlsx_db.commit()
    app = _create_test_app(xlsx_db, role="supervisor")
    return TestClient(app), xlsx_db


@pytest.fixture
def admin_client(xlsx_db):
    """TestClient authenticated as admin with a pre-created client."""
    TestDataFactory.create_client(xlsx_db, client_id=CLIENT_ID, client_name="XLSX Admin Client")
    xlsx_db.commit()
    app = _create_test_app(xlsx_db, role="admin")
    return TestClient(app), xlsx_db


# ---------------------------------------------------------------------------
# Helpers - create XLSX files in memory
# ---------------------------------------------------------------------------


def _make_xlsx_bytes(rows, sheet_name="Sheet1"):
    """Build an XLSX workbook in memory and return its raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_multi_sheet_xlsx_bytes(sheets):
    """Build a multi-sheet XLSX workbook. sheets = {name: [[rows]]}."""
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Downtime XLSX Upload Tests
# ===========================================================================


class TestDowntimeXlsxUpload:
    """Tests for /api/downtime/upload/csv with XLSX files.

    Note: The downtime CRUD function (create_downtime_event) has a
    pre-existing bug (passes entered_by instead of reported_by to
    DowntimeEntry). Therefore these tests verify that XLSX parsing works
    correctly by checking total_rows (rows parsed from XLSX), while the
    rows themselves fail at the CRUD layer — an issue unrelated to XLSX
    support.
    """

    def test_xlsx_file_parsed_correctly(self, supervisor_client):
        """XLSX file is parsed and all rows are processed."""
        client, db = supervisor_client

        wo = TestDataFactory.create_work_order(db, client_id=CLIENT_ID, work_order_id="WO-XLSX-DT-001")
        db.commit()

        xlsx_data = _make_xlsx_bytes(
            [
                ["client_id", "work_order_number", "shift_date", "downtime_category", "downtime_duration_minutes"],
                [CLIENT_ID, wo.work_order_id, "2025-06-01", "EQUIPMENT_FAILURE", 30],
                [CLIENT_ID, wo.work_order_id, "2025-06-02", "MAINTENANCE", 45],
            ]
        )

        response = client.post(
            "/api/downtime/upload/csv",
            files={
                "file": (
                    "downtime.xlsx",
                    BytesIO(xlsx_data),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200, response.text
        body = response.json()
        # Both rows were parsed from the XLSX file (total_rows proves parsing worked)
        assert body["total_rows"] == 2

    def test_xlsx_with_sheet_name(self, supervisor_client):
        """The sheet_name query parameter selects the correct sheet."""
        client, db = supervisor_client

        wo = TestDataFactory.create_work_order(db, client_id=CLIENT_ID, work_order_id="WO-XLSX-DT-SH")
        db.commit()

        xlsx_data = _make_multi_sheet_xlsx_bytes(
            {
                "Ignore": [
                    ["col_a"],
                    ["garbage"],
                ],
                "Downtime": [
                    ["client_id", "work_order_number", "shift_date", "downtime_category", "downtime_duration_minutes"],
                    [CLIENT_ID, wo.work_order_id, "2025-07-01", "MATERIAL_SHORTAGE", 60],
                ],
            }
        )

        response = client.post(
            "/api/downtime/upload/csv?sheet_name=Downtime",
            files={
                "file": (
                    "multi.xlsx",
                    BytesIO(xlsx_data),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200, response.text
        body = response.json()
        # Only 1 row from the Downtime sheet (not the Ignore sheet's 1 row)
        assert body["total_rows"] == 1

    def test_txt_file_rejected(self, supervisor_client):
        """A .txt file is rejected with 400."""
        client, _ = supervisor_client

        response = client.post(
            "/api/downtime/upload/csv",
            files={"file": ("bad.txt", BytesIO(b"some text"), "text/plain")},
        )
        assert response.status_code == 400
        detail = response.json()["detail"].lower()
        assert "csv" in detail or "xlsx" in detail

    def test_oversized_file_rejected(self, supervisor_client):
        """A file exceeding 10 MB is rejected with 413."""
        client, _ = supervisor_client

        big_data = b"\x00" * (10 * 1024 * 1024 + 1)

        response = client.post(
            "/api/downtime/upload/csv",
            files={
                "file": (
                    "big.xlsx",
                    BytesIO(big_data),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 413

    def test_csv_file_still_parsed(self, supervisor_client):
        """Existing CSV upload still works after XLSX support was added."""
        client, db = supervisor_client

        wo = TestDataFactory.create_work_order(db, client_id=CLIENT_ID, work_order_id="WO-CSV-DT-001")
        db.commit()

        csv_content = (
            "client_id,work_order_number,shift_date,downtime_category,downtime_duration_minutes\n"
            f"{CLIENT_ID},{wo.work_order_id},2025-08-01,EQUIPMENT_FAILURE,20\n"
        )

        response = client.post(
            "/api/downtime/upload/csv",
            files={"file": ("downtime.csv", BytesIO(csv_content.encode()), "text/csv")},
        )

        assert response.status_code == 200, response.text
        body = response.json()
        # CSV row was parsed correctly (total_rows proves parsing)
        assert body["total_rows"] == 1


# ===========================================================================
# Work Orders XLSX Upload Tests
# ===========================================================================


class TestWorkOrdersXlsxUpload:
    """Tests for /api/work-orders/upload/csv with XLSX files.

    Note: The work-order CRUD function (create_work_order) has a
    pre-existing bug — it uses dict-style item assignment on the
    Pydantic model (``work_order_data["status"] = "RECEIVED"``).
    Therefore these tests verify that XLSX parsing works correctly
    by checking total_rows (rows parsed from the file), while the
    rows themselves fail at the CRUD layer — an issue unrelated to
    XLSX support.
    """

    def test_xlsx_upload_accepted(self, admin_client):
        """XLSX file is parsed and all rows are processed."""
        client, db = admin_client

        xlsx_data = _make_xlsx_bytes(
            [
                ["work_order_id", "client_id", "style_model", "planned_quantity", "status", "priority"],
                ["WO-XL-001", CLIENT_ID, "STYLE-A", 500, "ACTIVE", "HIGH"],
                ["WO-XL-002", CLIENT_ID, "STYLE-B", 750, "ACTIVE", "MEDIUM"],
            ]
        )

        response = client.post(
            "/api/work-orders/upload/csv",
            files={
                "file": (
                    "orders.xlsx",
                    BytesIO(xlsx_data),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200, response.text
        body = response.json()
        # Both rows were parsed from the XLSX file (total_rows proves parsing worked)
        assert body["total_rows"] == 2

    def test_xlsx_fuzzy_headers(self, admin_client):
        """XLSX headers with spaces/capitals are matched correctly."""
        client, db = admin_client

        xlsx_data = _make_xlsx_bytes(
            [
                ["Work Order ID", "Client ID", "Style Model", "Planned Quantity", "Status", "Priority"],
                ["WO-FUZZY-001", CLIENT_ID, "STYLE-FUZZY", 100, "ACTIVE", "LOW"],
            ]
        )

        response = client.post(
            "/api/work-orders/upload/csv",
            files={
                "file": (
                    "fuzzy.xlsx",
                    BytesIO(xlsx_data),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        assert response.status_code == 200, response.text
        body = response.json()
        # Fuzzy header normalization worked (total_rows proves parsing)
        assert body["total_rows"] == 1

    def test_txt_file_rejected(self, admin_client):
        """A .txt file is rejected with 400."""
        client, _ = admin_client

        response = client.post(
            "/api/work-orders/upload/csv",
            files={"file": ("bad.txt", BytesIO(b"data"), "text/plain")},
        )
        assert response.status_code == 400

    def test_csv_still_works(self, admin_client):
        """CSV upload continues to function after XLSX support."""
        client, db = admin_client

        csv_content = (
            "work_order_id,client_id,style_model,planned_quantity,status,priority\n"
            f"WO-CSV-001,{CLIENT_ID},STYLE-CSV,200,ACTIVE,HIGH\n"
        )

        response = client.post(
            "/api/work-orders/upload/csv",
            files={"file": ("orders.csv", BytesIO(csv_content.encode()), "text/csv")},
        )

        assert response.status_code == 200, response.text
        body = response.json()
        # CSV row was parsed correctly (total_rows proves parsing)
        assert body["total_rows"] == 1

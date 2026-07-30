"""Characterization tests: Excel report sheets per report type (honest-surface PR)."""

from io import BytesIO

from openpyxl import load_workbook


def _sheets(test_client, auth_headers, report_type: str) -> list[str]:
    response = test_client.get(f"/api/reports/{report_type}/excel", headers=auth_headers)
    assert response.status_code == 200
    return list(load_workbook(BytesIO(response.content)).sheetnames)


class TestComprehensiveSheets:
    def test_comprehensive_has_all_data_sheets_and_no_charts(self, test_client, auth_headers):
        assert _sheets(test_client, auth_headers, "comprehensive") == [
            "Executive Summary",
            "Production Metrics",
            "Quality Metrics",
            "Downtime Analysis",
            "Attendance",
        ]


class TestPerTypeSheets:
    def test_production_excel_sheets(self, test_client, auth_headers):
        assert _sheets(test_client, auth_headers, "production") == [
            "Executive Summary",
            "Production Metrics",
            "Downtime Analysis",
        ]

    def test_quality_excel_sheets(self, test_client, auth_headers):
        assert _sheets(test_client, auth_headers, "quality") == [
            "Executive Summary",
            "Quality Metrics",
        ]

    def test_attendance_excel_sheets(self, test_client, auth_headers):
        assert _sheets(test_client, auth_headers, "attendance") == [
            "Executive Summary",
            "Attendance",
        ]

    def test_all_four_types_produce_distinct_sheet_sets(self, test_client, auth_headers):
        sheet_sets = {
            report_type: tuple(_sheets(test_client, auth_headers, report_type))
            for report_type in ("comprehensive", "production", "quality", "attendance")
        }
        assert len(set(sheet_sets.values())) == 4

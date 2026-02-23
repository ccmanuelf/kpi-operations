"""
Tests for backend.services.xlsx_parser

Validates XLSX-to-dict-rows parsing, header normalization,
cell type conversions, sheet selection, and edge cases.
All tests create workbooks in-memory using openpyxl (no disk I/O).
"""

import pytest
from datetime import datetime, date
from io import BytesIO

import openpyxl

from backend.services.xlsx_parser import (
    normalize_header,
    parse_xlsx_to_rows,
    get_sheet_names,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_xlsx(rows, sheet_name="Sheet1", header_row=None):
    """
    Create an in-memory XLSX workbook from a list of lists.

    The first sub-list is treated as the header row unless header_row is
    provided separately.

    Returns raw bytes suitable for parse_xlsx_to_rows.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _make_multi_sheet_xlsx(sheets_data):
    """
    Create a multi-sheet XLSX workbook.

    sheets_data: dict of {sheet_name: [[row1], [row2], ...]}
    Returns raw bytes.
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default = wb.active
    wb.remove(default)

    for name, rows in sheets_data.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ===========================================================================
# normalize_header tests
# ===========================================================================


class TestNormalizeHeader:
    """Tests for the normalize_header function."""

    def test_basic_normalization(self):
        assert normalize_header("Work Order ID") == "work_order_id"

    def test_strip_whitespace(self):
        assert normalize_header("  Shift Date  ") == "shift_date"

    def test_replace_hyphens(self):
        assert normalize_header("client-id") == "client_id"

    def test_replace_dots(self):
        assert normalize_header("qty.produced") == "qty_produced"

    def test_remove_special_chars(self):
        assert normalize_header("Units Produced (#)") == "units_produced"

    def test_collapse_multiple_underscores(self):
        assert normalize_header("work   order   id") == "work_order_id"

    def test_empty_string(self):
        assert normalize_header("") == ""

    def test_none_input(self):
        # None is falsy, so normalize_header returns empty string
        assert normalize_header(None) == ""

    def test_numeric_header(self):
        assert normalize_header("123") == "123"

    def test_mixed_case(self):
        assert normalize_header("Client ID") == "client_id"

    def test_already_normalized(self):
        assert normalize_header("work_order_id") == "work_order_id"


# ===========================================================================
# parse_xlsx_to_rows tests
# ===========================================================================


class TestParseXlsxToRows:
    """Tests for the parse_xlsx_to_rows function."""

    def test_basic_parsing(self):
        """Parse a simple 2-column, 2-row workbook."""
        content = _make_xlsx([
            ["Name", "Value"],
            ["Alpha", "100"],
            ["Beta", "200"],
        ])
        rows = parse_xlsx_to_rows(content)
        assert len(rows) == 2
        assert rows[0]["name"] == "Alpha"
        assert rows[0]["value"] == "100"
        assert rows[1]["name"] == "Beta"
        assert rows[1]["value"] == "200"

    def test_header_normalization(self):
        """Headers like 'Work Order ID' become 'work_order_id'."""
        content = _make_xlsx([
            ["Work Order ID", "Shift Date", "Client ID"],
            ["WO-001", "2025-01-10", "CLIENT-A"],
        ])
        rows = parse_xlsx_to_rows(content, fuzzy_headers=True)
        assert len(rows) == 1
        assert "work_order_id" in rows[0]
        assert "shift_date" in rows[0]
        assert "client_id" in rows[0]

    def test_fuzzy_headers_disabled(self):
        """With fuzzy_headers=False, headers are stripped but not normalized."""
        content = _make_xlsx([
            ["Work Order ID", "Shift Date"],
            ["WO-001", "2025-01-10"],
        ])
        rows = parse_xlsx_to_rows(content, fuzzy_headers=False)
        assert "Work Order ID" in rows[0]
        assert "Shift Date" in rows[0]

    def test_date_cell_date_only(self):
        """Date cells with midnight time become YYYY-MM-DD strings."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["date_col"])
        ws.append([datetime(2025, 3, 15, 0, 0, 0)])
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        rows = parse_xlsx_to_rows(content)
        assert rows[0]["date_col"] == "2025-03-15"

    def test_date_cell_with_time(self):
        """Date cells with non-midnight time include HH:MM:SS."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["datetime_col"])
        ws.append([datetime(2025, 3, 15, 14, 30, 45)])
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        rows = parse_xlsx_to_rows(content)
        assert rows[0]["datetime_col"] == "2025-03-15 14:30:45"

    def test_numeric_int_cell(self):
        """Integer cells are converted to string without decimal."""
        content = _make_xlsx([
            ["count"],
            [42],
        ])
        rows = parse_xlsx_to_rows(content)
        assert rows[0]["count"] == "42"

    def test_numeric_float_cell(self):
        """Float cells with fractional parts keep their decimals."""
        content = _make_xlsx([
            ["rate"],
            [3.14],
        ])
        rows = parse_xlsx_to_rows(content)
        assert rows[0]["rate"] == "3.14"

    def test_numeric_whole_float(self):
        """Float cells like 100.0 become '100' (no trailing .0)."""
        content = _make_xlsx([
            ["qty"],
            [100.0],
        ])
        rows = parse_xlsx_to_rows(content)
        assert rows[0]["qty"] == "100"

    def test_none_cell_becomes_empty_string(self):
        """None cell values become empty strings."""
        content = _make_xlsx([
            ["col_a", "col_b"],
            ["hello", None],
        ])
        rows = parse_xlsx_to_rows(content)
        assert rows[0]["col_b"] == ""

    def test_empty_rows_skipped(self):
        """Rows with all None values are skipped."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["col_a", "col_b"])
        ws.append(["data1", "data2"])
        ws.append([None, None])  # Empty row
        ws.append(["data3", "data4"])
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        rows = parse_xlsx_to_rows(content)
        assert len(rows) == 2
        assert rows[0]["col_a"] == "data1"
        assert rows[1]["col_a"] == "data3"

    def test_sheet_selection_by_name(self):
        """Selecting a specific sheet by name works."""
        content = _make_multi_sheet_xlsx({
            "Downtime": [
                ["event_id", "reason"],
                ["DT-001", "EQUIPMENT_FAILURE"],
            ],
            "Quality": [
                ["inspection_id", "result"],
                ["QI-001", "PASS"],
            ],
        })
        rows = parse_xlsx_to_rows(content, sheet_name="Quality")
        assert len(rows) == 1
        assert rows[0]["inspection_id"] == "QI-001"

    def test_sheet_not_found_raises(self):
        """Requesting a non-existent sheet raises ValueError."""
        content = _make_xlsx([["a"], [1]])
        with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
            parse_xlsx_to_rows(content, sheet_name="Missing")

    def test_empty_workbook(self):
        """A workbook with only headers and no data rows returns empty list."""
        content = _make_xlsx([
            ["col_a", "col_b"],
        ])
        rows = parse_xlsx_to_rows(content)
        assert rows == []

    def test_boolean_cell(self):
        """Boolean cells become '1' or '0'."""
        content = _make_xlsx([
            ["flag"],
            [True],
            [False],
        ])
        rows = parse_xlsx_to_rows(content)
        assert rows[0]["flag"] == "1"
        assert rows[1]["flag"] == "0"

    def test_invalid_file_raises(self):
        """Non-XLSX bytes raise ValueError."""
        with pytest.raises(ValueError, match="Cannot read XLSX file"):
            parse_xlsx_to_rows(b"this is not an xlsx file")

    def test_all_values_are_strings(self):
        """Every value in the returned dicts is a string."""
        content = _make_xlsx([
            ["text", "integer", "decimal", "empty"],
            ["hello", 42, 3.14, None],
        ])
        rows = parse_xlsx_to_rows(content)
        for key, value in rows[0].items():
            assert isinstance(value, str), f"Key '{key}' has non-string value: {type(value)}"

    def test_header_row_parameter(self):
        """Custom header_row parameter reads headers from a different row."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Metadata row - skip this"])
        ws.append(["col_x", "col_y"])  # Row 2 = headers
        ws.append(["val1", "val2"])
        buf = BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        rows = parse_xlsx_to_rows(content, header_row=2)
        assert len(rows) == 1
        assert rows[0]["col_x"] == "val1"


# ===========================================================================
# get_sheet_names tests
# ===========================================================================


class TestGetSheetNames:
    """Tests for the get_sheet_names function."""

    def test_single_sheet(self):
        content = _make_xlsx([["a"], [1]])
        names = get_sheet_names(content)
        assert names == ["Sheet1"]

    def test_multiple_sheets(self):
        content = _make_multi_sheet_xlsx({
            "Downtime": [["a"]],
            "Quality": [["b"]],
            "Attendance": [["c"]],
        })
        names = get_sheet_names(content)
        assert "Downtime" in names
        assert "Quality" in names
        assert "Attendance" in names
        assert len(names) == 3

    def test_invalid_file_raises(self):
        with pytest.raises(ValueError, match="Cannot read XLSX file"):
            get_sheet_names(b"not a real xlsx")

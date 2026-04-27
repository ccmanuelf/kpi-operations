"""
XLSX File Parser Service

Provides reusable XLSX-to-dict-rows parsing for upload endpoints.
Returns data in the same format as csv.DictReader (List[Dict[str, str]])
for backward compatibility with existing CSV processing code.
"""

import re
from datetime import date, datetime, time
from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl


def normalize_header(header: str) -> str:
    """
    Normalize a column header for fuzzy matching.

    Strips whitespace, lowercases, replaces spaces/hyphens with underscores,
    and removes special characters (keeping only alphanumeric and underscores).

    Args:
        header: Raw header string from the spreadsheet.

    Returns:
        Normalized header string (e.g., "Work Order ID" -> "work_order_id").
    """
    if not header:
        return ""
    h = str(header).strip().lower()
    # Replace spaces, hyphens, and dots with underscores
    h = re.sub(r"[\s\-\.]+", "_", h)
    # Remove anything that is not alphanumeric or underscore
    h = re.sub(r"[^a-z0-9_]", "", h)
    # Collapse multiple underscores
    h = re.sub(r"_+", "_", h)
    # Strip leading/trailing underscores
    h = h.strip("_")
    return h


def _cell_value_to_string(value: Any) -> str:
    """
    Convert an openpyxl cell value to a string, matching csv.DictReader behavior.

    - None -> ""
    - datetime -> "YYYY-MM-DD HH:MM:SS"
    - date -> "YYYY-MM-DD"
    - time -> "HH:MM:SS"
    - int/float -> string without trailing .0 for whole numbers
    - bool -> "1" / "0"
    - Everything else -> str()
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float):
        # Convert whole floats to int string (42.0 -> "42")
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value)


def get_sheet_names(file_content: bytes) -> List[str]:
    """
    Return list of sheet names in the workbook.

    Args:
        file_content: Raw bytes of the uploaded .xlsx file.

    Returns:
        List of sheet name strings.

    Raises:
        ValueError: If the file cannot be parsed as a valid XLSX workbook.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as exc:
        raise ValueError(f"Cannot read XLSX file: {exc}") from exc


def parse_xlsx_to_rows(
    file_content: bytes,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    fuzzy_headers: bool = True,
) -> List[Dict[str, str]]:
    """
    Parse an XLSX file into a list of dicts (same format as csv.DictReader).

    Args:
        file_content: Raw bytes of the uploaded .xlsx file.
        sheet_name: Sheet to read (default: first/active sheet).
        header_row: Row number containing headers (1-indexed, default 1).
        fuzzy_headers: If True, normalize headers (strip, lowercase, underscores).

    Returns:
        List of dicts where keys are column headers, values are cell values
        converted to strings. Empty rows are skipped.

    Raises:
        ValueError: If the file cannot be parsed or the requested sheet
            does not exist.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read XLSX file: {exc}") from exc

    try:
        # Select sheet
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. " f"Available sheets: {', '.join(wb.sheetnames)}")
            ws = wb[sheet_name]
        else:
            ws = wb.active
            if ws is None:
                # Fallback to first sheet if no active sheet is set
                if wb.sheetnames:
                    ws = wb[wb.sheetnames[0]]
                else:
                    raise ValueError("Workbook has no sheets")

        # Read header row
        raw_headers = []
        for cell in ws[header_row]:
            raw_headers.append(cell.value)

        # Build header list: normalize if fuzzy, otherwise keep as-is strings
        headers: List[str] = []
        for h in raw_headers:
            if h is None:
                headers.append("")
            elif fuzzy_headers:
                headers.append(normalize_header(str(h)))
            else:
                headers.append(str(h).strip())

        # Parse data rows (everything after header_row)
        rows: List[Dict[str, str]] = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            values = [cell.value for cell in row]

            # Skip completely empty rows
            if all(v is None for v in values):
                continue

            row_dict: Dict[str, str] = {}
            for col_idx, value in enumerate(values):
                if col_idx < len(headers) and headers[col_idx]:
                    row_dict[headers[col_idx]] = _cell_value_to_string(value)

            # Skip rows that are entirely empty strings after conversion
            if all(v == "" for v in row_dict.values()):
                continue

            rows.append(row_dict)

        return rows

    finally:
        wb.close()

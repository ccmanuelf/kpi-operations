"""
Excel Report Generator for KPI Platform
Generates formatted Excel reports with multiple sheets, charts, and formulas

IBM Carbon Design System color palette applied for consistent branding.
Reference: https://carbondesignsystem.com/guidelines/color/tokens
"""

from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Optional, List, Dict, Any, Sequence
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session

from backend.calculations.availability import calculate_availability_pure


class ExcelReportGenerator:
    """Generate comprehensive Excel reports for KPI data"""

    def __init__(self, db: Session):
        self.db = db

        # IBM Carbon Design System color palette
        # Reference: https://carbondesignsystem.com/guidelines/color/tokens
        self.colors = {
            # Primary colors
            "header": "FF0f62fe",  # IBM Blue 60 - Primary
            "header_dark": "FF0043ce",  # IBM Blue 70 - Primary dark
            "header_light": "FFedf5ff",  # IBM Blue 10 - Primary light bg
            # Status colors
            "success": "FF198038",  # Green 60 - Success
            "success_light": "FFdefbe6",  # Green 10 - Success background
            "warning": "FFf1c21b",  # Yellow 30 - Warning
            "warning_light": "FFfcf4d6",  # Yellow 10 - Warning background
            "error": "FFda1e28",  # Red 60 - Error/Danger
            "error_light": "FFfff1f1",  # Red 10 - Error background
            # Neutral colors
            "text_primary": "FF161616",  # Gray 100 - Primary text
            "text_secondary": "FF525252",  # Gray 70 - Secondary text
            "light_gray": "FFf4f4f4",  # Gray 10 - Layer/Background
            "medium_gray": "FFe0e0e0",  # Gray 20 - Border subtle
            "dark_gray": "FF525252",  # Gray 70 - Borders
        }

    def generate_report(
        self,
        client_id: Optional[str],
        start_date: date,
        end_date: date,
        output_path: Optional[Path] = None,
        sheets: Optional[Sequence[str]] = None,
    ) -> BytesIO:
        """
        Generate KPI Excel report.

        Args:
            client_id: Client ID (None for all clients)
            start_date: Report start date
            end_date: Report end date
            output_path: Optional file path to save Excel file
            sheets: Sheet keys to include ("summary", "production", "quality",
                "downtime", "attendance"). None = all sheets (comprehensive).

        Returns:
            BytesIO containing Excel data
        """
        sheet_builders = {
            "summary": self._create_summary_sheet,
            "production": self._create_production_sheet,
            "quality": self._create_quality_sheet,
            "downtime": self._create_downtime_sheet,
            "attendance": self._create_attendance_sheet,
        }
        selected = list(sheet_builders) if sheets is None else [key for key in sheet_builders if key in set(sheets)]

        if not selected:
            raise ValueError("sheets selected no valid sheet keys")

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        for key in selected:
            sheet_builders[key](wb, client_id, start_date, end_date)

        # Save to buffer or file
        buffer = BytesIO()
        wb.save(buffer if not output_path else str(output_path))
        buffer.seek(0)

        return buffer

    def _create_summary_sheet(self, wb: Workbook, client_id: Optional[str], start_date: date, end_date: date) -> None:
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")

        # Title
        ws["A1"] = "KPI Performance Dashboard"
        ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
        )
        ws.merge_cells("A1:F1")
        ws.row_dimensions[1].height = 30
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Metadata
        ws["A3"] = "Report Period:"
        ws["B3"] = f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
        ws["A4"] = "Generated:"
        ws["B4"] = datetime.now(tz=timezone.utc).strftime("%m/%d/%Y %I:%M %p")

        # Client info
        client_name = self._get_client_name(client_id)
        ws["A5"] = "Client:"
        ws["B5"] = client_name

        # KPI Summary Table Header
        ws["A7"] = "KPI"
        ws["B7"] = "Current Value"
        ws["C7"] = "Target"
        ws["D7"] = "Variance"
        ws["E7"] = "Status"
        ws["F7"] = "Trend"

        # Style header row
        for col in ["A", "B", "C", "D", "E", "F"]:
            cell = ws[f"{col}7"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch and populate KPI data
        kpi_data = self._fetch_kpi_summary_data(client_id, start_date, end_date)

        row = 8
        for kpi in kpi_data:
            ws[f"A{row}"] = kpi["name"]
            ws[f"B{row}"] = kpi["current"]
            ws[f"C{row}"] = kpi["target"]
            ws[f"D{row}"] = f"=B{row}-C{row}"
            ws[f"E{row}"] = kpi["status"]
            ws[f"F{row}"] = kpi["trend"]

            # Format values
            ws[f"B{row}"].number_format = kpi.get("format", "0.0")
            ws[f"C{row}"].number_format = kpi.get("format", "0.0")
            ws[f"D{row}"].number_format = kpi.get("format", "0.0")

            # Status color coding using Carbon Design tokens
            status_cell = ws[f"E{row}"]
            if kpi["status"] == "On Target":
                # Carbon Green 60 with white text
                status_cell.fill = PatternFill(
                    start_color=self.colors["success"], end_color=self.colors["success"], fill_type="solid"
                )
                status_cell.font = Font(bold=True, color="FFFFFF")
            elif kpi["status"] == "At Risk":
                # Carbon Yellow 30 with dark text (for contrast/accessibility)
                status_cell.fill = PatternFill(
                    start_color=self.colors["warning"], end_color=self.colors["warning"], fill_type="solid"
                )
                status_cell.font = Font(bold=True, color=self.colors["text_primary"][2:])  # Dark text for yellow bg
            elif kpi["status"] == "N/A":
                # Carbon Gray 20 with secondary text -- neutral/not-applicable,
                # distinct from the red error fallback below. Used by rows that
                # report a raw total rather than evaluating against a target
                # (Cycle 3 Task 4: Labor/OT Hours rows).
                status_cell.fill = PatternFill(
                    start_color=self.colors["medium_gray"], end_color=self.colors["medium_gray"], fill_type="solid"
                )
                status_cell.font = Font(bold=True, color=self.colors["text_secondary"][2:])
            else:
                # Carbon Red 60 with white text
                status_cell.fill = PatternFill(
                    start_color=self.colors["error"], end_color=self.colors["error"], fill_type="solid"
                )
                status_cell.font = Font(bold=True, color="FFFFFF")

            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Alternating row colors
            if row % 2 == 0:
                for col in ["A", "B", "C", "D", "F"]:
                    ws[f"{col}{row}"].fill = PatternFill(
                        start_color=self.colors["light_gray"], end_color=self.colors["light_gray"], fill_type="solid"
                    )

            row += 1

        # Apply borders
        self._apply_table_borders(ws, "A7", f"F{row-1}")

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

    def _create_production_sheet(
        self, wb: Workbook, client_id: Optional[str], start_date: date, end_date: date
    ) -> None:
        """Create production metrics sheet"""
        ws = wb.create_sheet("Production Metrics")

        # Header
        ws["A1"] = "Production Performance"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:H1")

        # Column headers
        headers = [
            "Date",
            "Product",
            "Units Produced",
            "Efficiency %",
            "Performance %",
            "Availability %",
            "OEE %",
            "Target OEE %",
        ]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch production data
        production_data = self._fetch_production_data(client_id, start_date, end_date)

        row = 4
        for entry in production_data:
            ws[f"A{row}"] = entry["date"]
            ws[f"B{row}"] = entry["product"]
            ws[f"C{row}"] = entry["units"]
            ws[f"D{row}"] = entry["efficiency"]
            ws[f"E{row}"] = entry["performance"]
            ws[f"F{row}"] = entry["availability"]
            ws[f"G{row}"] = f"=D{row}*E{row}*F{row}/10000"  # OEE formula
            ws[f"H{row}"] = 85

            # Number formatting
            for col in ["D", "E", "F", "G", "H"]:
                ws[f"{col}{row}"].number_format = '0.0"%"'

            row += 1

        # Add totals/averages
        ws[f"A{row}"] = "AVERAGE"
        ws[f"A{row}"].font = Font(bold=True)
        for col in ["D", "E", "F", "G"]:
            ws[f"{col}{row}"] = f"=AVERAGE({col}4:{col}{row-1})"
            ws[f"{col}{row}"].font = Font(bold=True)
            ws[f"{col}{row}"].number_format = '0.0"%"'

        self._apply_table_borders(ws, "A3", f"H{row}")

        # Adjust column widths
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col].width = 15

    def _create_quality_sheet(self, wb: Workbook, client_id: Optional[str], start_date: date, end_date: date) -> None:
        """Create quality metrics sheet"""
        ws = wb.create_sheet("Quality Metrics")

        # Header
        ws["A1"] = "Quality Performance"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:G1")

        # Column headers
        headers = ["Date", "Product", "Units Inspected", "Units Passed", "FPY %", "PPM", "DPMO"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch quality data
        quality_data = self._fetch_quality_data(client_id, start_date, end_date)

        row = 4
        for entry in quality_data:
            ws[f"A{row}"] = entry["date"]
            ws[f"B{row}"] = entry["product"]
            ws[f"C{row}"] = entry["inspected"]
            ws[f"D{row}"] = entry["passed"]
            ws[f"E{row}"] = f"=D{row}/C{row}*100"  # FPY formula
            ws[f"F{row}"] = f"=(C{row}-D{row})/C{row}*1000000"  # PPM formula
            ws[f"G{row}"] = entry.get("dpmo", 0)

            # Number formatting
            ws[f"E{row}"].number_format = '0.00"%"'
            ws[f"F{row}"].number_format = "#,##0"
            ws[f"G{row}"].number_format = "#,##0"

            row += 1

        # Add averages
        ws[f"A{row}"] = "AVERAGE"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"E{row}"] = f"=AVERAGE(E4:E{row-1})"
        ws[f"F{row}"] = f"=AVERAGE(F4:F{row-1})"
        ws[f"G{row}"] = f"=AVERAGE(G4:G{row-1})"

        for col in ["E", "F", "G"]:
            ws[f"{col}{row}"].font = Font(bold=True)

        self._apply_table_borders(ws, "A3", f"G{row}")

        # Adjust column widths
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 18

    @staticmethod
    def _downtime_category_label(category_code: str) -> str:
        """Display label for a root_cause_category enum code (spec §7): no
        backend i18n, so title-case the underscore-free code, e.g.
        'machine' -> 'Machine', 'uncategorized' -> 'Uncategorized'."""
        return category_code.replace("_", " ").title()

    def _create_downtime_sheet(self, wb: Workbook, client_id: Optional[str], start_date: date, end_date: date) -> None:
        """Create downtime analysis sheet"""
        ws = wb.create_sheet("Downtime Analysis")

        # Header
        ws["A1"] = "Downtime Events"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")

        # Fetch downtime data
        downtime_data = self._fetch_downtime_data(client_id, start_date, end_date)

        # By-category summary (Cycle 1 minimal rollup)
        from collections import defaultdict

        totals: dict[str, dict[str, float]] = defaultdict(lambda: {"events": 0, "minutes": 0.0})
        for entry in downtime_data:
            cat = entry["category"] if entry["category"] != "—" else "uncategorized"
            totals[cat]["events"] += 1
            totals[cat]["minutes"] += entry["duration"] * 60
        grand = sum(v["minutes"] for v in totals.values()) or 1.0

        summary_header_row = 3
        summary_headers = ["By Category", "Events", "Total Minutes", "% of Total"]
        for idx, header in enumerate(summary_headers, start=1):
            cell = ws.cell(row=summary_header_row, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        summary_row = summary_header_row + 1
        for cat in sorted(totals):
            ws.cell(row=summary_row, column=1, value=self._downtime_category_label(cat))
            ws.cell(row=summary_row, column=2, value=totals[cat]["events"])
            ws.cell(row=summary_row, column=3, value=round(totals[cat]["minutes"], 1))
            ws.cell(row=summary_row, column=4, value=round(100.0 * totals[cat]["minutes"] / grand, 1))
            summary_row += 1

        if summary_row > summary_header_row + 1:
            self._apply_table_borders(ws, f"A{summary_header_row}", f"D{summary_row - 1}")

        # Column headers (detail table) — starts one blank row below the summary block
        detail_header_row = summary_row + 1
        headers = ["Date", "Machine/Line", "Category", "Duration (hrs)", "Impact %", "Root Cause"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=detail_header_row, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = detail_header_row + 1
        for entry in downtime_data:
            ws[f"A{row}"] = entry["date"]
            ws[f"B{row}"] = entry["machine"]
            ws[f"C{row}"] = entry["category"]
            ws[f"D{row}"] = entry["duration"]
            ws[f"E{row}"] = entry["impact"]
            ws[f"F{row}"] = entry.get("root_cause", "N/A")

            ws[f"D{row}"].number_format = "0.00"
            ws[f"E{row}"].number_format = '0.0"%"'

            row += 1

        # Add totals
        ws[f"A{row}"] = "TOTAL"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"D{row}"] = f"=SUM(D{detail_header_row + 1}:D{row - 1})"
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"D{row}"].number_format = "0.00"

        self._apply_table_borders(ws, f"A{detail_header_row}", f"F{row}")

        # Adjust column widths
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions["F"].width = 30

    def _create_attendance_sheet(
        self, wb: Workbook, client_id: Optional[str], start_date: date, end_date: date
    ) -> None:
        """Create attendance/absenteeism sheet"""
        ws = wb.create_sheet("Attendance")

        # Header
        ws["A1"] = "Attendance & Absenteeism"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:E1")

        # Column headers
        headers = ["Date", "Scheduled Employees", "Absent", "Present", "Absenteeism Rate %"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.colors["header"], end_color=self.colors["header"], fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Fetch attendance data
        attendance_data = self._fetch_attendance_data(client_id, start_date, end_date)

        row = 4
        for entry in attendance_data:
            ws[f"A{row}"] = entry["date"]
            ws[f"B{row}"] = entry["scheduled"]
            ws[f"C{row}"] = entry["absent"]
            ws[f"D{row}"] = f"=B{row}-C{row}"
            ws[f"E{row}"] = f"=C{row}/B{row}*100"

            ws[f"E{row}"].number_format = '0.0"%"'

            row += 1

        # Add averages
        ws[f"A{row}"] = "AVERAGE"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"E{row}"] = f"=AVERAGE(E4:E{row-1})"
        ws[f"E{row}"].font = Font(bold=True)
        ws[f"E{row}"].number_format = '0.0"%"'

        self._apply_table_borders(ws, "A3", f"E{row}")

        # Adjust column widths
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 20

    def _apply_table_borders(self, ws: Any, start_cell: str, end_cell: str) -> None:
        """Apply borders to table range"""
        thin_border = Border(
            left=Side(style="thin", color=self.colors["dark_gray"]),
            right=Side(style="thin", color=self.colors["dark_gray"]),
            top=Side(style="thin", color=self.colors["dark_gray"]),
            bottom=Side(style="thin", color=self.colors["dark_gray"]),
        )

        start_row = ws[start_cell].row
        start_col = ws[start_cell].column
        end_row = ws[end_cell].row
        end_col = ws[end_cell].column

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border

    def _get_client_name(self, client_id: Optional[str]) -> str:
        """Get client name from database"""
        if not client_id:
            return "All Clients"

        from backend.orm.client import Client

        client = self.db.query(Client).filter(Client.client_id == client_id).first()
        return client.client_name if client else "Unknown Client"

    def _fetch_kpi_summary_data(
        self, client_id: Optional[str], start_date: date, end_date: date
    ) -> List[Dict[str, Any]]:
        """Fetch KPI summary data from database"""
        from backend.orm.production_entry import ProductionEntry
        from backend.orm.quality_entry import QualityEntry
        from backend.orm.attendance_entry import AttendanceEntry
        from backend.orm.product import Product

        kpi_data = []

        # Production KPIs
        production_query = self.db.query(ProductionEntry).filter(
            ProductionEntry.production_date.between(
                datetime.combine(start_date, datetime.min.time()), datetime.combine(end_date, datetime.max.time())
            )
        )

        if client_id:
            production_query = production_query.join(Product).filter(Product.client_id == client_id)

        production_entries = production_query.all()

        if production_entries:
            # Efficiency
            total_efficiency = sum(float(e.efficiency_percentage or 0) for e in production_entries)
            avg_efficiency = total_efficiency / len(production_entries)
            kpi_data.append(
                {
                    "name": "Efficiency",
                    "current": avg_efficiency,
                    "target": 85,
                    "status": "On Target" if avg_efficiency >= 85 else "At Risk",
                    "trend": "↑" if avg_efficiency >= 85 else "↓",
                    "format": '0.0"%"',
                }
            )

            # Performance
            total_performance = sum(float(e.performance_percentage or 0) for e in production_entries)
            avg_performance = total_performance / len(production_entries)
            kpi_data.append(
                {
                    "name": "Performance",
                    "current": avg_performance,
                    "target": 85,
                    "status": "On Target" if avg_performance >= 85 else "At Risk",
                    "trend": "↑" if avg_performance >= 85 else "→",
                    "format": '0.0"%"',
                }
            )

        # Quality KPIs
        quality_query = self.db.query(QualityEntry).filter(
            QualityEntry.shift_date.between(
                datetime.combine(start_date, datetime.min.time()), datetime.combine(end_date, datetime.max.time())
            )
        )

        if client_id:
            quality_query = quality_query.filter(QualityEntry.client_id == client_id)

        quality_entries = quality_query.all()

        if quality_entries:
            total_inspected = sum(e.units_inspected for e in quality_entries)
            total_defects = sum(e.units_defective for e in quality_entries)

            # FPY
            fpy = ((total_inspected - total_defects) / total_inspected * 100) if total_inspected > 0 else 0
            kpi_data.append(
                {
                    "name": "FPY",
                    "current": fpy,
                    "target": 99,
                    "status": "On Target" if fpy >= 99 else "At Risk",
                    "trend": "→",
                    "format": '0.0"%"',
                }
            )

            # PPM
            ppm = (total_defects / total_inspected * 1_000_000) if total_inspected > 0 else 0
            kpi_data.append(
                {
                    "name": "PPM",
                    "current": ppm,
                    "target": 1000,
                    "status": "On Target" if ppm <= 1000 else "At Risk",
                    "trend": "↑" if ppm > 1000 else "↓",
                    "format": "#,##0",
                }
            )

        # Attendance KPIs
        attendance_query = self.db.query(AttendanceEntry).filter(
            AttendanceEntry.shift_date.between(
                datetime.combine(start_date, datetime.min.time()), datetime.combine(end_date, datetime.max.time())
            )
        )
        if client_id:
            attendance_query = attendance_query.filter(AttendanceEntry.client_id == client_id)

        attendance_entries = attendance_query.all()

        if attendance_entries:
            total_scheduled = sum(float(e.scheduled_hours or 0) for e in attendance_entries)
            total_absent = sum(float(e.scheduled_hours or 0) for e in attendance_entries if e.is_absent == 1)
            absenteeism = (total_absent / total_scheduled * 100) if total_scheduled > 0 else 0

            kpi_data.append(
                {
                    "name": "Absenteeism",
                    "current": absenteeism,
                    "target": 5,
                    "status": "On Target" if absenteeism <= 5 else "At Risk",
                    "trend": "↓" if absenteeism <= 5 else "↑",
                    "format": '0.0"%"',
                }
            )

        # On-Time Delivery (justified-delay-flag Cycle 2, Task 7): gross +
        # net-of-justified, as two rows (this table has one KPI per row, no
        # per-KPI secondary column). Sourced from calculate_true_otd's
        # standard_otd mode — "all orders with delivery dates", the closest
        # analog to this table's other broad averages (true_otd restricts to
        # COMPLETED-only orders, a narrower definition than the rest of this
        # sheet). Both rows come from the SAME calculate_true_otd(...) call so
        # gross and net stay internally consistent.
        #
        # Only rendered for a single-client report: calculate_true_otd is
        # inherently single-client (mirrors /api/kpi/otd's true_otd/
        # standard_otd keys, which are likewise omitted for multi/all-client
        # scope) — an "all clients" comprehensive report has no single-client
        # rollup to compute here.
        if client_id:
            from backend.calculations.otd import calculate_true_otd

            otd_result = calculate_true_otd(self.db, client_id, start_date, end_date)
            # Guard: with zero delivered orders in the window, standard_otd's
            # percentage/net_percentage are both 0 (no data, not "0% OTD") —
            # rendering them would show a misleading "0.0% / At Risk" row.
            if otd_result["standard_otd"]["total"] > 0:
                otd_gross = float(otd_result["standard_otd"]["percentage"])
                otd_net = float(otd_result["standard_otd"]["net_percentage"])
                kpi_data.append(
                    {
                        "name": "OTD",
                        "current": otd_gross,
                        "target": 95,
                        "status": "On Target" if otd_gross >= 95 else "At Risk",
                        "trend": "→",
                        "format": '0.0"%"',
                    }
                )
                kpi_data.append(
                    {
                        "name": "OTD (Net of Justified)",
                        "current": otd_net,
                        "target": 95,
                        "status": "On Target" if otd_net >= 95 else "At Risk",
                        "trend": "→",
                        "format": '0.0"%"',
                    }
                )

        # Labor Hours (Cycle 3 Task 4): billed / available-for-efficiency
        # totals plus OT double/triple tiers, sourced from a single
        # summarize_labor_hours (Task 1) call -- same single-data-source-call,
        # client-scoped-list, zero-data-guard idiom as the OTD block above.
        #
        # Only rendered for a single-client report, matching the OTD
        # precedent and /api/kpi/labor-hours (which also never resolves an
        # unbounded all-clients read via this path).
        #
        # These are raw hour totals, not %-of-target KPIs like every other
        # row in this table -- there is no meaningful target to evaluate
        # them against. The status-color loop above only distinguishes
        # "On Target" (green) / "At Risk" (yellow) and falls back to red for
        # any other value, so a naive status here would paint an
        # uninformative total as a false alarm. We reuse this file's
        # existing "N/A" not-applicable convention (already used for missing
        # text fields elsewhere in this module, e.g. root_cause/machine
        # columns) as the status value, paired with a matching neutral
        # (gray) branch added to the status-color loop above. Target is left
        # None (blank cell) for the same reason -- no evaluation is made, so
        # no target value is meaningful.
        if client_id:
            from backend.calculations.labor_hours import summarize_labor_hours

            labor_result = summarize_labor_hours(self.db, [client_id], start_date, end_date)
            # Guard: zero attendance entries in the window is an
            # absence-of-data signal (not "0.0 hours worked") -- omit the
            # rows entirely, mirroring the OTD zero-delivered guard above.
            if labor_result["entry_counts"]["total"] > 0:
                totals = labor_result["totals"]
                for label, value in (
                    ("Labor Hours — Actual", totals["actual"]),
                    ("Labor Hours — Billed", totals["billed"]),
                    ("Labor Hours — Available", totals["available_for_efficiency"]),
                    ("OT Hours — Double", totals["double"]),
                    ("OT Hours — Triple", totals["triple"]),
                ):
                    kpi_data.append(
                        {
                            "name": label,
                            "current": float(value),
                            "target": None,
                            "status": "N/A",
                            "trend": "→",
                            "format": "0.00",
                        }
                    )

        return kpi_data

    def _fetch_production_data(
        self, client_id: Optional[str], start_date: date, end_date: date
    ) -> List[Dict[str, Any]]:
        """Fetch production data from database"""
        from backend.orm.production_entry import ProductionEntry
        from backend.orm.product import Product
        from sqlalchemy import func

        query = (
            self.db.query(
                ProductionEntry.production_date,
                Product.product_name,
                func.sum(ProductionEntry.units_produced).label("units"),
                func.avg(ProductionEntry.efficiency_percentage).label("efficiency"),
                func.avg(ProductionEntry.performance_percentage).label("performance"),
                func.sum(ProductionEntry.run_time_hours).label("run_hours"),
                func.sum(ProductionEntry.downtime_hours).label("downtime_hours"),
            )
            .join(Product)
            .filter(
                ProductionEntry.production_date.between(
                    datetime.combine(start_date, datetime.min.time()),
                    datetime.combine(end_date, datetime.max.time()),
                )
            )
        )

        if client_id:
            query = query.filter(Product.client_id == client_id)

        query = query.group_by(ProductionEntry.production_date, Product.product_name).order_by(
            ProductionEntry.production_date
        )

        results = query.all()

        return [
            {
                "date": r.production_date,
                "product": r.product_name,
                "units": r.units or 0,
                "efficiency": float(r.efficiency or 0),
                "performance": float(r.performance or 0),
                "availability": float(
                    calculate_availability_pure(
                        Decimal(str((r.run_hours or 0))) + Decimal(str(r.downtime_hours or 0)),
                        Decimal(str(r.downtime_hours or 0)),
                    )
                ),
            }
            for r in results
        ]

    def _fetch_quality_data(self, client_id: Optional[str], start_date: date, end_date: date) -> List[Dict[str, Any]]:
        """Fetch quality data from database"""
        from backend.orm.quality_entry import QualityEntry

        # QualityEntry is keyed by shift_date and carries client_id directly
        # (no Product join needed). `inspection_date` is optional metadata.
        query = self.db.query(QualityEntry).filter(
            QualityEntry.shift_date.between(
                datetime.combine(start_date, datetime.min.time()), datetime.combine(end_date, datetime.max.time())
            )
        )

        if client_id:
            query = query.filter(QualityEntry.client_id == client_id)

        results = query.order_by(QualityEntry.shift_date).all()

        return [
            {
                "date": r.inspection_date or r.shift_date,
                "product": "—",  # QualityEntry tracks WO not product directly
                "inspected": r.units_inspected,
                "passed": (r.units_inspected or 0) - (r.units_defective or 0),
                "dpmo": float(r.dpmo or 0),
            }
            for r in results
        ]

    def _fetch_downtime_data(self, client_id: Optional[str], start_date: date, end_date: date) -> List[Dict[str, Any]]:
        """Fetch downtime data from database"""
        from backend.orm.downtime_entry import DowntimeEntry

        query = self.db.query(DowntimeEntry).filter(
            DowntimeEntry.shift_date.between(
                datetime.combine(start_date, datetime.min.time()), datetime.combine(end_date, datetime.max.time())
            )
        )

        if client_id:
            query = query.filter(DowntimeEntry.client_id == client_id)

        results = query.order_by(DowntimeEntry.shift_date).all()

        return [
            {
                "date": r.shift_date,
                "machine": r.machine_id or "N/A",
                "category": r.root_cause_category or "—",
                "duration": float((r.downtime_duration_minutes or 0) / 60),  # hours
                "impact": 0.0,  # Calculate based on scheduled time
                "root_cause": r.downtime_reason,
            }
            for r in results
        ]

    def _fetch_attendance_data(
        self, client_id: Optional[str], start_date: date, end_date: date
    ) -> List[Dict[str, Any]]:
        """Fetch attendance data from database"""
        from backend.orm.attendance_entry import AttendanceEntry
        from sqlalchemy import func, case

        query = self.db.query(
            AttendanceEntry.shift_date,
            func.count(AttendanceEntry.attendance_entry_id).label("scheduled"),
            func.sum(case((AttendanceEntry.is_absent == 1, 1), else_=0)).label("absent"),
        ).filter(
            AttendanceEntry.shift_date.between(
                datetime.combine(start_date, datetime.min.time()), datetime.combine(end_date, datetime.max.time())
            )
        )
        if client_id:
            query = query.filter(AttendanceEntry.client_id == client_id)
        query = query.group_by(AttendanceEntry.shift_date).order_by(AttendanceEntry.shift_date)

        results = query.all()

        return [{"date": r.shift_date, "scheduled": r.scheduled or 0, "absent": r.absent or 0} for r in results]
